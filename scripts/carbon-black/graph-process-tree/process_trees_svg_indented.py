#!/usr/bin/env python3
import argparse
import csv
import hashlib
import os
import re
import textwrap
import zipfile
from collections import defaultdict
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook

DEDUP_COLS = [
    "hostname",
    "parent_name",
    "parent_pid",
    "process_name",
    "cmdline",
    "username",
    "process_pid",
]

ROOT_PARENT_VALUES = {"", "-1", "0", "None", "nan", "null"}


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value}".rstrip("0").rstrip(".")
    return str(value).strip()


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def safe_slug(value: str, limit: int = 60) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    cleaned = cleaned.strip("._-") or "unknown"
    return cleaned[:limit]


def dedupe_key(row):
    return tuple(normalize_value(row.get(col, "")) for col in DEDUP_COLS)


def make_node_id(row):
    return hashlib.sha1("||".join(dedupe_key(row)).encode("utf-8")).hexdigest()


def read_rows(input_xlsx: str, sheet_ref):
    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    if isinstance(sheet_ref, int):
        ws = wb[wb.sheetnames[sheet_ref]]
    else:
        ws = wb[sheet_ref]

    row_iter = ws.iter_rows(values_only=True)
    headers = [normalize_value(h) for h in next(row_iter)]
    idx = {name: i for i, name in enumerate(headers)}

    missing = [c for c in DEDUP_COLS if c not in idx]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")

    raw_rows = []
    for values in row_iter:
        row = {c: normalize_value(values[idx[c]]) for c in DEDUP_COLS}
        raw_rows.append(row)
    return ws.title, raw_rows


def dedupe_rows(raw_rows):
    seen = set()
    unique_rows = []
    for row in raw_rows:
        key = dedupe_key(row)
        if key in seen:
            continue
        seen.add(key)
        record = {c: row[c] for c in DEDUP_COLS}
        record["node_id"] = make_node_id(record)
        unique_rows.append(record)
    return unique_rows


def choose_parent(child, candidates):
    filtered = [c for c in candidates if c["node_id"] != child["node_id"]]
    if not filtered:
        return None

    parent_name = child.get("parent_name", "")
    if parent_name:
        exact_name = [c for c in filtered if c.get("process_name", "") == parent_name]
        if exact_name:
            filtered = exact_name

    username = child.get("username", "")
    if len(filtered) > 1 and username:
        same_user = [c for c in filtered if c.get("username", "") == username]
        if same_user:
            filtered = same_user

    filtered = sorted(
        filtered,
        key=lambda c: (
            c.get("process_name", ""),
            c.get("username", ""),
            c.get("cmdline", ""),
            c.get("hostname", ""),
            c.get("node_id", ""),
        ),
    )
    return filtered[0]


def build_forest(unique_rows):
    node_map = {}
    by_host_pid = defaultdict(list)
    for row in unique_rows:
        node_map[row["node_id"]] = row
        by_host_pid[(row["hostname"], row["process_pid"])] .append(row)

    children = defaultdict(list)
    parent_of = {}
    roots = []
    edge_count = 0

    for row in unique_rows:
        parent_pid = row.get("parent_pid", "")
        if parent_pid in ROOT_PARENT_VALUES:
            roots.append(row["node_id"])
            continue

        candidates = by_host_pid.get((row["hostname"], parent_pid), [])
        parent = choose_parent(row, candidates)
        if not parent:
            roots.append(row["node_id"])
            continue

        parent_of[row["node_id"]] = parent["node_id"]
        children[parent["node_id"]].append(row["node_id"])
        edge_count += 1

    for node_id in children:
        children[node_id] = sorted(
            children[node_id],
            key=lambda cid: (
                node_map[cid].get("process_name", ""),
                node_map[cid].get("process_pid", ""),
                node_map[cid].get("username", ""),
            ),
        )

    return node_map, children, parent_of, roots, edge_count


def collect_component(root_id, children, assigned):
    ordered = []
    edges = []
    depths = {}

    def dfs(node_id, depth):
        if node_id in assigned:
            return
        assigned.add(node_id)
        depths[node_id] = depth
        ordered.append(node_id)
        for child_id in children.get(node_id, []):
            edges.append((node_id, child_id))
            dfs(child_id, depth + 1)

    dfs(root_id, 0)
    return ordered, edges, depths


def split_components(node_map, children, roots):
    assigned = set()
    components = []

    for root_id in roots:
        if root_id in assigned:
            continue
        ordered, edges, depths = collect_component(root_id, children, assigned)
        components.append({
            "root_id": root_id,
            "nodes": ordered,
            "edges": edges,
            "depths": depths,
            "is_rooted": True,
        })

    for node_id in sorted(node_map):
        if node_id in assigned:
            continue
        ordered, edges, depths = collect_component(node_id, children, assigned)
        components.append({
            "root_id": node_id,
            "nodes": ordered,
            "edges": edges,
            "depths": depths,
            "is_rooted": False,
        })

    components.sort(key=lambda c: (-len(c["nodes"]), node_map[c["root_id"]]["hostname"], node_map[c["root_id"]]["process_name"], node_map[c["root_id"]]["process_pid"]))
    return components


def node_lines(row, cmdline_wrap_width, cmdline_max_chars):
    hostname = row.get("hostname", "") or "unknown-host"
    process_name = row.get("process_name", "") or "unknown-process"
    username = row.get("username", "") or "unknown-user"
    pid = row.get("process_pid", "") or "?"
    parent_pid = row.get("parent_pid", "") or "-"
    parent_name = row.get("parent_name", "") or "root"
    cmdline = normalize_space(row.get("cmdline", "")) or "(empty)"
    if cmdline_max_chars and len(cmdline) > cmdline_max_chars:
        cmdline = cmdline[: cmdline_max_chars - 3] + "..."

    body = [
        f"user: {username}",
        f"host: {hostname}",
        f"pid: {pid}    ppid: {parent_pid}",
        f"parent: {parent_name}",
    ]
    cmd_prefix = "cmd: "
    cmd_wrapped = textwrap.wrap(
        cmd_prefix + cmdline,
        width=cmdline_wrap_width,
        break_long_words=False,
        break_on_hyphens=False,
    ) or [cmd_prefix + "(empty)"]
    return process_name, body, cmd_wrapped


def measure_node(row, cmdline_wrap_width, cmdline_max_chars, box_width):
    title, body, cmd = node_lines(row, cmdline_wrap_width, cmdline_max_chars)
    header_h = 28
    line_h = 17
    top_pad = 12
    bottom_pad = 12
    height = top_pad + header_h + (len(body) * line_h) + (len(cmd) * line_h) + bottom_pad
    return {
        "title": title,
        "body": body,
        "cmd": cmd,
        "w": box_width,
        "h": height,
        "header_h": header_h,
        "line_h": line_h,
        "top_pad": top_pad,
    }


def render_svg(component, node_map, out_path: Path, box_width: int, depth_step: int, v_gap: int, margin_x: int, margin_y: int, cmdline_wrap_width: int, cmdline_max_chars: int):
    layout = {}
    current_y = margin_y + 56
    max_depth = 0

    for node_id in component["nodes"]:
        row = node_map[node_id]
        node = measure_node(row, cmdline_wrap_width, cmdline_max_chars, box_width)
        depth = component["depths"][node_id]
        x = margin_x + depth * depth_step
        y = current_y
        node.update({"x": x, "y": y, "depth": depth})
        layout[node_id] = node
        current_y += node["h"] + v_gap
        max_depth = max(max_depth, depth)

    canvas_width = margin_x * 2 + box_width + max_depth * depth_step
    canvas_height = current_y + margin_y

    root = node_map[component["root_id"]]
    title = f"Root process tree - host={root['hostname']} | root={root['process_name']} | pid={root['process_pid']} | nodes={len(component['nodes'])}"

    parts = []
    parts.append('<?xml version="1.0" encoding="UTF-8"?>')
    parts.append(f'<svg xmlns="http://www.w3.org/2000/svg" width="{canvas_width}" height="{canvas_height}" viewBox="0 0 {canvas_width} {canvas_height}">')
    parts.append('<defs>')
    parts.append('<style><![CDATA[')
    parts.append('.title{font:700 22px Helvetica,Arial,sans-serif; fill:#0f172a;}')
    parts.append('.subtitle{font:400 12px Helvetica,Arial,sans-serif; fill:#475569;}')
    parts.append('.nodeTitle{font:700 13px Helvetica,Arial,sans-serif; fill:#0f172a;}')
    parts.append('.nodeBody{font:400 12px Helvetica,Arial,sans-serif; fill:#334155;}')
    parts.append('.nodeCmd{font:400 11px "Courier New",monospace; fill:#1e293b;}')
    parts.append('.procBadge{font:700 10px Helvetica,Arial,sans-serif; fill:#1d4ed8; letter-spacing:1px;}')
    parts.append(']]></style>')
    parts.append('</defs>')
    parts.append('<rect x="0" y="0" width="100%" height="100%" fill="#f8fafc"/>')
    parts.append(f'<text x="{margin_x}" y="32" class="title">{escape(title)}</text>')
    parts.append(f'<text x="{margin_x}" y="50" class="subtitle">Deduped by hostname, parent_name, parent_pid, process_name, cmdline, username, process_pid</text>')

    # edges first
    for parent_id, child_id in component["edges"]:
        p = layout[parent_id]
        c = layout[child_id]
        start_x = p["x"] + p["w"]
        start_y = p["y"] + p["h"] / 2
        elbow_x = start_x + 34
        end_x = c["x"]
        end_y = c["y"] + c["h"] / 2
        parts.append(
            f'<path d="M {start_x:.1f} {start_y:.1f} L {elbow_x:.1f} {start_y:.1f} L {elbow_x:.1f} {end_y:.1f} L {end_x:.1f} {end_y:.1f}" '
            f'fill="none" stroke="#94a3b8" stroke-width="1.5"/>'
        )
        parts.append(f'<circle cx="{end_x:.1f}" cy="{end_y:.1f}" r="2.4" fill="#64748b"/>')

    # nodes
    for node_id in component["nodes"]:
        row = node_map[node_id]
        node = layout[node_id]
        x = node["x"]
        y = node["y"]
        w = node["w"]
        h = node["h"]
        header_h = node["header_h"]
        parts.append(f'<g id="{node_id}">')
        parts.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="14" ry="14" fill="#ffffff" stroke="#53779b" stroke-width="1.6"/>')
        parts.append(f'<rect x="{x}" y="{y}" width="{w}" height="{header_h + 10}" rx="14" ry="14" fill="#dbeafe"/>')
        parts.append(f'<rect x="{x}" y="{y + header_h}" width="{w}" height="10" fill="#dbeafe"/>')
        parts.append(f'<text x="{x + 14}" y="{y + 22}" class="nodeTitle">{escape(node["title"])}</text>')
        parts.append(f'<text x="{x + w - 48}" y="{y + 21}" class="procBadge">PROC</text>')

        text_y = y + header_h + 16
        for line in node["body"]:
            parts.append(f'<text x="{x + 14}" y="{text_y}" class="nodeBody">{escape(line)}</text>')
            text_y += node["line_h"]
        for line in node["cmd"]:
            parts.append(f'<text x="{x + 14}" y="{text_y}" class="nodeCmd">{escape(line)}</text>')
            text_y += node["line_h"]
        parts.append('</g>')

    parts.append('</svg>')
    out_path.write_text("\n".join(parts), encoding="utf-8")


def write_csv(path: Path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def make_zip(zip_path: Path, base_dir: Path):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in sorted(base_dir.rglob("*")):
            if file_path.is_file() and file_path != zip_path:
                zf.write(file_path, arcname=file_path.relative_to(base_dir))


def parse_args():
    p = argparse.ArgumentParser(description="Generate one SVG per root process tree from an Excel workbook.")
    p.add_argument("input_xlsx", help="Path to the input workbook")
    p.add_argument("--sheet", default="0", help="Worksheet index or name (default: first sheet)")
    p.add_argument("--outdir", default="process_tree_svgs_indented", help="Output directory")
    p.add_argument("--box-width", type=int, default=460, help="Node card width in px")
    p.add_argument("--depth-step", type=int, default=560, help="Horizontal step between generations in px")
    p.add_argument("--v-gap", type=int, default=26, help="Vertical gap between nodes in px")
    p.add_argument("--cmdline-wrap-width", type=int, default=54, help="Approximate character wrap for cmdline")
    p.add_argument("--cmdline-max-chars", type=int, default=0, help="Optional cmdline truncation; 0 keeps full text")
    return p.parse_args()


def main():
    args = parse_args()
    sheet_ref = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
    outdir = Path(args.outdir)
    trees_dir = outdir / "trees"
    trees_dir.mkdir(parents=True, exist_ok=True)

    sheet_name, raw_rows = read_rows(args.input_xlsx, sheet_ref)
    total_rows = len(raw_rows)
    unique_rows = dedupe_rows(raw_rows)
    node_map, children, parent_of, roots, edge_count = build_forest(unique_rows)
    components = split_components(node_map, children, roots)

    manifest_rows = []
    for idx, component in enumerate(components, start=1):
        root = node_map[component["root_id"]]
        filename = f"{idx:03d}__{safe_slug(root['hostname'])}__{safe_slug(root['process_name'])}__pid{safe_slug(root['process_pid'])}.svg"
        out_path = trees_dir / filename
        render_svg(
            component=component,
            node_map=node_map,
            out_path=out_path,
            box_width=args.box_width,
            depth_step=args.depth_step,
            v_gap=args.v_gap,
            margin_x=24,
            margin_y=24,
            cmdline_wrap_width=args.cmdline_wrap_width,
            cmdline_max_chars=args.cmdline_max_chars,
        )
        manifest_rows.append({
            "tree_index": idx,
            "file": os.path.relpath(out_path, outdir),
            "hostname": root["hostname"],
            "root_process_name": root["process_name"],
            "root_process_pid": root["process_pid"],
            "username": root["username"],
            "node_count": len(component["nodes"]),
            "edge_count": len(component["edges"]),
            "max_depth": max(component["depths"].values()) if component["depths"] else 0,
            "root_status": "rooted" if component["is_rooted"] else "unrooted_fragment",
        })

    deduped_csv = outdir / "process_nodes_deduped.csv"
    write_csv(deduped_csv, [{col: row[col] for col in DEDUP_COLS} for row in unique_rows], DEDUP_COLS)

    manifest_csv = outdir / "tree_manifest.csv"
    write_csv(
        manifest_csv,
        manifest_rows,
        [
            "tree_index",
            "file",
            "hostname",
            "root_process_name",
            "root_process_pid",
            "username",
            "node_count",
            "edge_count",
            "max_depth",
            "root_status",
        ],
    )

    summary_txt = outdir / "summary.txt"
    largest = manifest_rows[0] if manifest_rows else None
    lines = [
        f"Workbook sheet: {sheet_name}",
        f"Input rows: {total_rows}",
        f"Unique rows after dedupe: {len(unique_rows)}",
        f"Duplicates removed: {total_rows - len(unique_rows)}",
        f"Edges created: {edge_count}",
        f"Trees rendered: {len(manifest_rows)}",
        f"Root/unmatched nodes: {len(roots)}",
        f"Diagram directory: {trees_dir}",
        f"Tree manifest: {manifest_csv}",
        f"Deduped CSV: {deduped_csv}",
    ]
    if largest:
        lines.extend([
            f"Largest tree index: {largest['tree_index']}",
            f"Largest tree file: {largest['file']}",
            f"Largest tree nodes: {largest['node_count']}",
            f"Largest tree max depth: {largest['max_depth']}",
        ])
    summary_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")

    zip_path = outdir / "process_tree_svgs_bundle.zip"
    make_zip(zip_path, outdir)

    print(f"[OK] Wrote {len(manifest_rows)} SVG files to {trees_dir}")
    print(f"[OK] Wrote {manifest_csv}")
    print(f"[OK] Wrote {deduped_csv}")
    print(f"[OK] Wrote {summary_txt}")
    print(f"[OK] Wrote {zip_path}")


if __name__ == "__main__":
    main()
