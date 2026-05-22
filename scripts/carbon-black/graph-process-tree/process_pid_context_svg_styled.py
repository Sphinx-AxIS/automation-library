#!/usr/bin/env python3
"""
Generate a PID-centered process diagram using the same SVG card style as the
root-tree renderer.

Behavior
- Reads an Excel workbook (.xlsx)
- Deduplicates rows using:
    hostname, parent_name, parent_pid, process_name, cmdline, username, process_pid
- Finds rows whose process_pid == target PID (optionally filtered by hostname)
- Walks upward from the target node to the root parent
- Walks downward from the target node to include all descendants
- Writes one styled SVG per matching target row
- Also writes a CSV for each rendered subgraph plus a manifest.csv

Dependency
    pip install openpyxl
"""

import argparse
import csv
import hashlib
import os
import re
import textwrap
from collections import defaultdict, deque
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook

DEDUP_COLS = [
    "hostname",
    "parent_name",
    "parent_pid",
    "process_name",
    "cmdline",
    "username",
    "process_pid",
]

ROOT_PARENT_VALUES = {"", "-1", "0", "None", "nan", "null"}


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value}".rstrip("0").rstrip(".")
    return str(value).strip()


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_pid_text(value: str) -> str:
    text = normalize_value(value)
    if not text:
        return ""
    try:
        return str(int(float(text)))
    except Exception:
        return text


def safe_slug(value: str, limit: int = 60) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value).strip())
    cleaned = cleaned.strip("._-") or "unknown"
    return cleaned[:limit]


def dedupe_key(row):
    return tuple(normalize_value(row.get(col, "")) for col in DEDUP_COLS)


def make_node_id(row):
    return hashlib.sha1("||".join(dedupe_key(row)).encode("utf-8")).hexdigest()


def read_rows(input_xlsx: str, sheet_ref):
    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    if isinstance(sheet_ref, int):
        ws = wb[wb.sheetnames[sheet_ref]]
    else:
        ws = wb[sheet_ref]

    row_iter = ws.iter_rows(values_only=True)
    headers = [normalize_value(h) for h in next(row_iter)]
    idx = {name: i for i, name in enumerate(headers)}

    missing = [c for c in DEDUP_COLS if c not in idx]
    if missing:
        raise ValueError(f"Missing expected columns: {missing}")

    raw_rows = []
    for values in row_iter:
        row = {c: normalize_value(values[idx[c]]) for c in DEDUP_COLS}
        raw_rows.append(row)
    return ws.title, raw_rows


def dedupe_rows(raw_rows):
    seen = set()
    unique_rows = []
    for row in raw_rows:
        key = dedupe_key(row)
        if key in seen:
            continue
        seen.add(key)
        record = {c: row[c] for c in DEDUP_COLS}
        record["process_pid_norm"] = normalize_pid_text(record["process_pid"])
        record["parent_pid_norm"] = normalize_pid_text(record["parent_pid"])
        record["node_id"] = make_node_id(record)
        unique_rows.append(record)
    return unique_rows


def choose_parent(child, candidates):
    filtered = [c for c in candidates if c["node_id"] != child["node_id"]]
    if not filtered:
        return None

    parent_name = child.get("parent_name", "")
    if parent_name:
        exact_name = [c for c in filtered if c.get("process_name", "") == parent_name]
        if exact_name:
            filtered = exact_name

    username = child.get("username", "")
    if len(filtered) > 1 and username:
        same_user = [c for c in filtered if c.get("username", "") == username]
        if same_user:
            filtered = same_user

    filtered = sorted(
        filtered,
        key=lambda c: (
            c.get("process_name", ""),
            c.get("username", ""),
            c.get("cmdline", ""),
            c.get("hostname", ""),
            c.get("node_id", ""),
        ),
    )
    return filtered[0]


def build_indexes(unique_rows):
    node_map = {}
    by_host_pid = defaultdict(list)
    by_host_parentpid = defaultdict(list)

    for row in unique_rows:
        node_map[row["node_id"]] = row
        by_host_pid[(row["hostname"], row["process_pid_norm"])].append(row)
        by_host_parentpid[(row["hostname"], row["parent_pid_norm"])].append(row)

    return node_map, by_host_pid, by_host_parentpid


def get_parent(node, by_host_pid):
    parent_pid = node.get("parent_pid_norm", "")
    if parent_pid in ROOT_PARENT_VALUES:
        return None
    candidates = by_host_pid.get((node["hostname"], parent_pid), [])
    return choose_parent(node, candidates)


def get_children(node, by_host_parentpid):
    pid = node.get("process_pid_norm", "")
    if not pid:
        return []
    candidates = by_host_parentpid.get((node["hostname"], pid), [])

    process_name = node.get("process_name", "")
    exact = [c for c in candidates if c.get("parent_name", "") == process_name]
    other = [c for c in candidates if c.get("parent_name", "") != process_name]
    ordered = sorted(
        exact + other,
        key=lambda c: (
            c.get("process_name", ""),
            c.get("process_pid_norm", ""),
            c.get("username", ""),
            c.get("node_id", ""),
        ),
    )
    return ordered


def collect_pid_context(start_node_id, node_map, by_host_pid, by_host_parentpid):
    included_nodes = set()
    edges = set()
    ancestors = set()
    descendants = set()

    lineage = []
    current = node_map[start_node_id]
    seen_up = set()

    while current and current["node_id"] not in seen_up:
        current_id = current["node_id"]
        seen_up.add(current_id)
        included_nodes.add(current_id)
        lineage.append(current_id)
        if current_id != start_node_id:
            ancestors.add(current_id)

        parent = get_parent(current, by_host_pid)
        if parent:
            included_nodes.add(parent["node_id"])
            edges.add((parent["node_id"], current_id))
        current = parent

    queue = deque([start_node_id])
    seen_down = {start_node_id}

    while queue:
        node_id = queue.popleft()
        node = node_map[node_id]
        for child in get_children(node, by_host_parentpid):
            child_id = child["node_id"]
            included_nodes.add(child_id)
            edges.add((node_id, child_id))
            if child_id != start_node_id:
                descendants.add(child_id)
            if child_id not in seen_down:
                seen_down.add(child_id)
                queue.append(child_id)

    depths = {}
    for depth, node_id in enumerate(reversed(lineage)):
        depths[node_id] = depth

    queue = deque([start_node_id])
    while queue:
        parent_id = queue.popleft()
        parent_depth = depths.get(parent_id, 0)
        for src, dst in sorted(edges):
            if src == parent_id and dst not in depths:
                depths[dst] = parent_depth + 1
                queue.append(dst)

    ordered_nodes = sorted(
        included_nodes,
        key=lambda nid: (
            depths.get(nid, 0),
            node_map[nid].get("process_name", ""),
            node_map[nid].get("process_pid_norm", ""),
            node_map[nid].get("username", ""),
            node_map[nid].get("node_id", ""),
        ),
    )

    return {
        "target_id": start_node_id,
        "nodes": ordered_nodes,
        "edges": sorted(edges),
        "depths": depths,
        "ancestors": ancestors,
        "descendants": descendants,
        "lineage": lineage,
        "root_id": lineage[-1] if lineage else start_node_id,
    }


def node_lines(row, cmdline_wrap_width, cmdline_max_chars):
    hostname = row.get("hostname", "") or "unknown-host"
    process_name = row.get("process_name", "") or "unknown-process"
    username = row.get("username", "") or "unknown-user"
    pid = row.get("process_pid", "") or "?"
    parent_pid = row.get("parent_pid", "") or "-"
    parent_name = row.get("parent_name", "") or "root"
    cmdline = normalize_space(row.get("cmdline", "")) or "(empty)"
    if cmdline_max_chars and len(cmdline) > cmdline_max_chars:
        cmdline = cmdline[: cmdline_max_chars - 3] + "..."

    body = [
        f"user: {username}",
        f"host: {hostname}",
        f"pid: {pid}    ppid: {parent_pid}",
        f"parent: {parent_name}",
    ]
    cmd_prefix = "cmd: "
    cmd_wrapped = textwrap.wrap(
        cmd_prefix + cmdline,
        width=cmdline_wrap_width,
        break_long_words=False,
        break_on_hyphens=False,
    ) or [cmd_prefix + "(empty)"]
    return process_name, body, cmd_wrapped


def measure_node(row, cmdline_wrap_width, cmdline_max_chars, box_width):
    title, body, cmd = node_lines(row, cmdline_wrap_width, cmdline_max_chars)
    header_h = 28
    line_h = 17
    top_pad = 12
    bottom_pad = 12
    height = top_pad + header_h + (len(body) * line_h) + (len(cmd) * line_h) + bottom_pad
    return {
        "title": title,
        "body": body,
        "cmd": cmd,
        "w": box_width,
        "h": height,
        "header_h": header_h,
        "line_h": line_h,
        "top_pad": top_pad,
    }


def render_svg(component, node_map, out_path, box_width, depth_step, v_gap, margin_x, margin_y, cmdline_wrap_width, cmdline_max_chars):
    layout = {}
    current_y = margin_y + 70
    max_depth = 0

    for node_id in component["nodes"]:
        row = node_map[node_id]
        node = measure_node(row, cmdline_wrap_width, cmdline_max_chars, box_width)
        depth = component["depths"].get(node_id, 0)
        x = margin_x + depth * depth_step
        y = current_y
        node.update({"x": x, "y": y, "depth": depth})
        layout[node_id] = node
        current_y += node["h"] + v_gap
        max_depth = max(max_depth, depth)

    canvas_width = margin_x * 2 + box_width + max_depth * depth_step + 40
    canvas_height = current_y + margin_y

    target = node_map[component["target_id"]]
    root = node_map[component["root_id"]]
    title = (
        f"PID context - host={target['hostname']} | target={target['process_name']} | "
        f"pid={target['process_pid']} | root={root['process_name']}:{root['process_pid']} | "
        f"nodes={len(component['nodes'])}"
    )

    parts = []
    parts.append('<?xml version="1.0" encoding="UTF-8"?>')
    parts.append(f'<svg xmlns="http://www.w3.org/2000/svg" width="{canvas_width}" height="{canvas_height}" viewBox="0 0 {canvas_width} {canvas_height}">')
    parts.append('<defs>')
    parts.append('<style><![CDATA[')
    parts.append('.title{font:700 22px Helvetica,Arial,sans-serif; fill:#0f172a;}')
    parts.append('.subtitle{font:400 12px Helvetica,Arial,sans-serif; fill:#475569;}')
    parts.append('.nodeTitle{font:700 13px Helvetica,Arial,sans-serif; fill:#0f172a;}')
    parts.append('.nodeBody{font:400 12px Helvetica,Arial,sans-serif; fill:#334155;}')
    parts.append('.nodeCmd{font:400 11px "Courier New",monospace; fill:#1e293b;}')
    parts.append('.procBadge{font:700 10px Helvetica,Arial,sans-serif; fill:#1d4ed8; letter-spacing:1px;}')
    parts.append(']]></style>')
    parts.append('</defs>')
    parts.append('<rect x="0" y="0" width="100%" height="100%" fill="#f8fafc"/>')
    parts.append(f'<text x="{margin_x}" y="32" class="title">{escape(title)}</text>')
    parts.append(f'<text x="{margin_x}" y="50" class="subtitle">Deduped by hostname, parent_name, parent_pid, process_name, cmdline, username, process_pid</text>')

    for parent_id, child_id in component["edges"]:
        p = layout[parent_id]
        c = layout[child_id]
        start_x = p["x"] + p["w"]
        start_y = p["y"] + p["h"] / 2
        elbow_x = start_x + 34
        end_x = c["x"]
        end_y = c["y"] + c["h"] / 2
        parts.append(
            f'<path d="M {start_x:.1f} {start_y:.1f} L {elbow_x:.1f} {start_y:.1f} L {elbow_x:.1f} {end_y:.1f} L {end_x:.1f} {end_y:.1f}" '
            f'fill="none" stroke="#94a3b8" stroke-width="1.5"/>'
        )
        parts.append(f'<circle cx="{end_x:.1f}" cy="{end_y:.1f}" r="2.4" fill="#64748b"/>')

    for node_id in component["nodes"]:
        row = node_map[node_id]
        node = layout[node_id]
        x = node["x"]
        y = node["y"]
        w = node["w"]
        h = node["h"]
        header_h = node["header_h"]

        if node_id == component["target_id"]:
            border = "#dc2626"
            stroke_width = "2.2"
            header_fill = "#fee2e2"
            badge_text = "TARGET"
            badge_fill = "#b91c1c"
        elif node_id in component["ancestors"]:
            border = "#53779b"
            stroke_width = "1.6"
            header_fill = "#dbeafe"
            badge_text = "PROC"
            badge_fill = "#1d4ed8"
        elif node_id in component["descendants"]:
            border = "#53779b"
            stroke_width = "1.6"
            header_fill = "#e2fbe8"
            badge_text = "CHILD"
            badge_fill = "#15803d"
        else:
            border = "#53779b"
            stroke_width = "1.6"
            header_fill = "#dbeafe"
            badge_text = "PROC"
            badge_fill = "#1d4ed8"

        parts.append(f'<g id="{node_id}">')
        parts.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="14" ry="14" fill="#ffffff" stroke="{border}" stroke-width="{stroke_width}"/>')
        parts.append(f'<rect x="{x}" y="{y}" width="{w}" height="{header_h + 10}" rx="14" ry="14" fill="{header_fill}"/>')
        parts.append(f'<rect x="{x}" y="{y + header_h}" width="{w}" height="10" fill="{header_fill}"/>')
        parts.append(f'<text x="{x + 14}" y="{y + 22}" class="nodeTitle">{escape(node["title"])}</text>')
        parts.append(f'<text x="{x + w - 62}" y="{y + 21}" class="procBadge" fill="{badge_fill}">{badge_text}</text>')

        text_y = y + header_h + 16
        for line in node["body"]:
            parts.append(f'<text x="{x + 14}" y="{text_y}" class="nodeBody">{escape(line)}</text>')
            text_y += node["line_h"]
        for line in node["cmd"]:
            parts.append(f'<text x="{x + 14}" y="{text_y}" class="nodeCmd">{escape(line)}</text>')
            text_y += node["line_h"]
        parts.append('</g>')

    parts.append('</svg>')
    out_path.write_text("\n".join(parts), encoding="utf-8")


def write_csv(path: Path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def parse_args():
    p = argparse.ArgumentParser(description="Generate a PID-centered SVG using the same card layout as the root-tree renderer.")
    p.add_argument("input_xlsx", help="Path to the input workbook")
    p.add_argument("--pid", required=True, help="Target process PID")
    p.add_argument("--hostname", default="", help="Optional hostname filter")
    p.add_argument("--sheet", default="0", help="Worksheet index or name (default: first sheet)")
    p.add_argument("--outdir", default="pid_context_svg", help="Output directory")
    p.add_argument("--box-width", type=int, default=460, help="Node card width in px")
    p.add_argument("--depth-step", type=int, default=560, help="Horizontal step between generations in px")
    p.add_argument("--v-gap", type=int, default=26, help="Vertical gap between nodes in px")
    p.add_argument("--cmdline-wrap-width", type=int, default=54, help="Approximate character wrap for cmdline")
    p.add_argument("--cmdline-max-chars", type=int, default=0, help="Optional cmdline truncation; 0 keeps full text")
    return p.parse_args()


def main():
    args = parse_args()
    sheet_ref = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    sheet_name, raw_rows = read_rows(args.input_xlsx, sheet_ref)
    unique_rows = dedupe_rows(raw_rows)
    node_map, by_host_pid, by_host_parentpid = build_indexes(unique_rows)

    target_pid = normalize_pid_text(args.pid)
    matches = []
    for row in unique_rows:
        if row["process_pid_norm"] != target_pid:
            continue
        if args.hostname and row["hostname"].lower() != args.hostname.lower():
            continue
        matches.append(row)

    if not matches:
        raise SystemExit(
            f"No deduplicated rows found for process_pid={target_pid}"
            + (f" and hostname={args.hostname}" if args.hostname else "")
        )

    manifest_rows = []
    for idx, row in enumerate(matches, start=1):
        component = collect_pid_context(row["node_id"], node_map, by_host_pid, by_host_parentpid)
        filename = (
            f"{idx:03d}__{safe_slug(row['hostname'])}__"
            f"{safe_slug(row['process_name'])}__pid{safe_slug(row['process_pid'])}.svg"
        )
        out_path = outdir / filename
        render_svg(
            component=component,
            node_map=node_map,
            out_path=out_path,
            box_width=args.box_width,
            depth_step=args.depth_step,
            v_gap=args.v_gap,
            margin_x=24,
            margin_y=24,
            cmdline_wrap_width=args.cmdline_wrap_width,
            cmdline_max_chars=args.cmdline_max_chars,
        )

        csv_rows = []
        for node_id in component["nodes"]:
            item = {col: node_map[node_id][col] for col in DEDUP_COLS}
            item["role"] = (
                "target" if node_id == component["target_id"]
                else "ancestor" if node_id in component["ancestors"]
                else "descendant" if node_id in component["descendants"]
                else "other"
            )
            csv_rows.append(item)

        csv_path = outdir / f"{Path(filename).stem}.csv"
        write_csv(csv_path, csv_rows, DEDUP_COLS + ["role"])

        manifest_rows.append({
            "match_index": idx,
            "svg_file": os.path.relpath(out_path, outdir),
            "csv_file": os.path.relpath(csv_path, outdir),
            "hostname": row["hostname"],
            "target_process_name": row["process_name"],
            "target_process_pid": row["process_pid"],
            "target_username": row["username"],
            "root_process_name": node_map[component["root_id"]]["process_name"],
            "root_process_pid": node_map[component["root_id"]]["process_pid"],
            "node_count": len(component["nodes"]),
            "edge_count": len(component["edges"]),
            "max_depth": max(component["depths"].values()) if component["depths"] else 0,
        })

    manifest_path = outdir / "manifest.csv"
    write_csv(
        manifest_path,
        manifest_rows,
        [
            "match_index",
            "svg_file",
            "csv_file",
            "hostname",
            "target_process_name",
            "target_process_pid",
            "target_username",
            "root_process_name",
            "root_process_pid",
            "node_count",
            "edge_count",
            "max_depth",
        ],
    )

    print(f"Sheet: {sheet_name}")
    print(f"Input rows: {len(raw_rows)}")
    print(f"Unique rows after dedupe: {len(unique_rows)}")
    print(f"PID matches rendered: {len(matches)}")
    print(f"Wrote output to: {outdir}")


if __name__ == "__main__":
    main()
